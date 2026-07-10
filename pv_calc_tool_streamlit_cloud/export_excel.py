from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def _dict_to_df(data: dict) -> pd.DataFrame:
    return pd.DataFrame([{"指标": key, "值": value} for key, value in data.items()])


def _build_core_summary(project_info: dict, capacity: dict, consumption: dict, cost: dict, revenue_summary: dict, suggestions: list[str]) -> pd.DataFrame:
    return _dict_to_df(
        {
            "项目名称": project_info.get("项目名称", ""),
            "项目地点": project_info.get("项目地点", ""),
            "装机容量(kWp)": capacity.get("dc_capacity_kwp", ""),
            "装机容量(MWp)": (capacity.get("dc_capacity_kwp", 0) or 0) / 1000,
            "有效安装面积(㎡)": capacity.get("effective_install_area_m2", ""),
            "年用电量(kWh)": consumption.get("annual_consumption_kwh", ""),
            "首年发电量(kWh)": revenue_summary.get("首年发电量(kWh)", consumption.get("pv_generation_kwh", "")),
            "自发自用比例": consumption.get("self_use_ratio", ""),
            "余电上网比例": consumption.get("feed_in_ratio", ""),
            "总投资(元)": cost.get("total_investment_yuan", revenue_summary.get("总投资(元)", "")),
            "总投资(万元)": (cost.get("total_investment_yuan", revenue_summary.get("总投资(元)", 0)) or 0) / 10000,
            "单瓦投资(元/Wp)": revenue_summary.get("单瓦投资(元/Wp)", ""),
            "年均收入(万元)": (revenue_summary.get("年均收入(元)", 0) or 0) / 10000,
            "年均净现金流(万元)": (revenue_summary.get("年均净现金流(元)", 0) or 0) / 10000,
            "静态回收期(年)": revenue_summary.get("静态回收期(年)", ""),
            "动态回收期(年)": revenue_summary.get("动态回收期(年)", ""),
            "IRR": revenue_summary.get("项目IRR", ""),
            "NPV(万元)": (revenue_summary.get("NPV(元)", 0) or 0) / 10000,
            "LCOE(元/kWh)": revenue_summary.get("LCOE(元/kWh)", ""),
            "主要风险提示": "；".join(suggestions),
            "测算口径说明": revenue_summary.get("收益测算口径说明", ""),
        }
    )


COST_ITEM_LABELS = {
    "module": "组件",
    "inverter": "逆变器",
    "mounting": "支架",
    "cable_tray": "电缆桥架",
    "distribution": "配电设备",
    "grid_connection": "并网接入",
    "civil_reinforcement": "土建及加固",
    "installation": "施工安装",
    "design_supervision_testing": "设计/监理/检测/验收",
    "project_management": "项目管理及其他",
}


def _format_workbook(writer):
    workbook = writer.book
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in sheet.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 12), 36)
        if sheet.title == "敏感性分析":
            headers = {cell.value: cell.column for cell in sheet[1]}
            for header in ["变化幅度", "IRR"]:
                col_idx = headers.get(header)
                if col_idx:
                    for cell in sheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                        for item in cell:
                            item.number_format = "0.00%"


def build_excel_report(
    project_info: dict,
    capacity: dict,
    consumption: dict,
    cost: dict,
    revenue_summary: dict,
    cashflow_table: pd.DataFrame,
    sensitivity_table: pd.DataFrame,
    suggestions: list[str],
    holiday_calendar_table: pd.DataFrame | None = None,
    cost_unit_costs: dict | None = None,
    grid_analysis: dict | None = None,
    grid_scenario_table: pd.DataFrame | None = None,
) -> bytes:
    """Excel 导出：生成包含基础参数、五步测算、现金流、敏感性和风险建议的工作簿。"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _build_core_summary(project_info, capacity, consumption, cost, revenue_summary, suggestions).to_excel(
            writer, sheet_name="核心结论摘要", index=False
        )
        _dict_to_df(project_info).to_excel(writer, sheet_name="项目基础参数", index=False)
        capacity_copy = dict(capacity)
        capacity_copy["公式说明"] = "有效安装面积 = 输入面积 × 实际应用屋顶利用率 × (1 - 实际应用综合扣减比例) × (1 - 修正后组件排布损耗系数)"
        _dict_to_df(capacity_copy).to_excel(writer, sheet_name="容量测算", index=False)
        _dict_to_df(consumption).to_excel(writer, sheet_name="消纳测算", index=False)
        cost_copy = dict(cost)
        cost_items = dict(cost_copy.get("cost_items_yuan", {}) or {})
        cost_summary = {key: value for key, value in cost_copy.items() if key != "cost_items_yuan"}
        _dict_to_df(cost_summary).to_excel(writer, sheet_name="造价测算", index=False)
        if cost_items:
            unit_costs = cost_unit_costs or {}
            detail = pd.DataFrame(
                [
                    {
                        "分项名称": COST_ITEM_LABELS.get(key, key),
                        "单价元/Wp": unit_costs.get(key, ""),
                        "金额元": value,
                        "金额万元": value / 10000,
                        "备注": "",
                    }
                    for key, value in cost_items.items()
                ]
            )
            detail.to_excel(writer, sheet_name="造价分项明细", index=False)
        _dict_to_df(revenue_summary).to_excel(writer, sheet_name="收益测算", index=False)
        if grid_analysis:
            _dict_to_df(grid_analysis).to_excel(writer, sheet_name="并网方式分析", index=False)
        if grid_scenario_table is not None and not grid_scenario_table.empty:
            grid_scenario_table.to_excel(writer, sheet_name="并网容量方案推荐", index=False)
        cashflow_table.to_excel(writer, sheet_name="25年现金流", index=False)
        sensitivity_export = sensitivity_table.copy()
        if "NPV(元)" in sensitivity_export.columns:
            sensitivity_export["NPV(万元)"] = sensitivity_export["NPV(元)"] / 10000
        sensitivity_export.to_excel(writer, sheet_name="敏感性分析", index=False)
        if holiday_calendar_table is not None and not holiday_calendar_table.empty:
            holiday_calendar_table.to_excel(writer, sheet_name="节假日与停产修正", index=False)
        pd.DataFrame({"风险提示与优化建议": suggestions}).to_excel(writer, sheet_name="风险提示与优化建议", index=False)
        _format_workbook(writer)
    return output.getvalue()
