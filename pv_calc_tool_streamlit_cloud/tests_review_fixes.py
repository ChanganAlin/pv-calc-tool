from __future__ import annotations

from io import BytesIO
from datetime import date, datetime

import pandas as pd
from openpyxl import load_workbook

from modules.bill_parser import parse_excel_bill
from modules.capacity_calc import calculate_capacity
from modules.consumption_calc import (
    calculate_precise_consumption,
    calculate_monthly_tou_consumption,
    normalize_month_value,
    normalize_monthly_tou_table,
    normalize_power_or_energy_curve,
)
from modules.export_excel import build_excel_report
from modules.holiday_calendar import build_holiday_calendar_config, load_holiday_config
from modules.image_area_estimator import estimate_area_from_pixel_area
from modules.revenue_calc import calculate_revenue


def test_capacity_area_basis():
    total = calculate_capacity(1000, 2, 0.6, 0.8, 0.1, 0.05, 1.1, area_basis="roof_total_area")
    usable = calculate_capacity(1000, 2, 0.6, 0.8, 0.1, 0.05, 1.1, area_basis="usable_area")
    confirmed = calculate_capacity(1000, 2, 0.6, 0.8, 0.1, 0.05, 1.1, area_basis="installable_area")
    assert total.effective_install_area_m2 == 684
    assert usable.effective_install_area_m2 == 855
    assert confirmed.effective_install_area_m2 == 950
    assert confirmed.applied_roof_utilization_ratio == 1
    assert confirmed.applied_deduction_ratio == 0


def test_normalize_month_value_and_table():
    assert normalize_month_value(1) == 1
    assert normalize_month_value("1月") == 1
    assert normalize_month_value("2026-01") == 1
    assert normalize_month_value("2026/01") == 1
    assert normalize_month_value("2026年1月") == 1
    assert normalize_month_value("2026年01月") == 1
    assert normalize_month_value("2026-01-01") == 1
    assert normalize_month_value("2026/01/01") == 1
    assert normalize_month_value(pd.Timestamp("2026-01-01")) == 1
    assert normalize_month_value(date(2026, 1, 1)) == 1
    assert normalize_month_value(datetime(2026, 1, 1, 12, 0)) == 1
    assert normalize_month_value("2026/12/05") == 12
    df = pd.DataFrame(
        {
            "月份": ["2026-01", "1月", "2026年2月"],
            "总电量": [100, 200, 300],
            "尖电量": [10, 20, 30],
            "峰电量": [20, 30, 40],
            "平电量": [30, 40, 50],
            "谷电量": [40, 50, 60],
        }
    )
    normalized = normalize_monthly_tou_table(df)
    jan = normalized.loc[normalized["月份"] == 1].iloc[0]
    assert jan["总电量(kWh)"] == 300
    assert list(normalized["月份"]) == [1, 2]


def test_scale_area_differs_by_pixel_length():
    area_56 = estimate_area_from_pixel_area(10000, 10, 56)
    area_100 = estimate_area_from_pixel_area(10000, 10, 100)
    assert area_56 > area_100


def test_precise_curve_resample_and_tolerance():
    load_raw = pd.DataFrame({"time": pd.date_range("2026-01-01", periods=8, freq="15min"), "load_kw": [100] * 8})
    pv_raw = pd.DataFrame({"time": pd.date_range("2026-01-01", periods=2, freq="60min"), "pv_kwh": [80, 80]})
    load = normalize_power_or_energy_curve(load_raw)
    pv = normalize_power_or_energy_curve(pv_raw, value_kind="pv")
    result, detail = calculate_precise_consumption(load, 160, pv, 100)
    assert result.self_use_kwh > 0
    assert detail["pv_kwh"].sum() == 160
    assert len(detail) == 8

    shifted_pv_raw = pd.DataFrame({"time": pd.date_range("2026-01-01 00:20", periods=8, freq="15min"), "pv_kwh": [10] * 8})
    shifted_pv = normalize_power_or_energy_curve(shifted_pv_raw, value_kind="pv")
    shifted_result, shifted_detail = calculate_precise_consumption(load, 80, shifted_pv, 100)
    assert shifted_detail["pv_kwh"].sum() < 80
    assert "未匹配点" in shifted_result.data_quality_note


def test_monthly_tou_calendar_merge_with_text_month():
    holiday_df = load_holiday_config("data/holiday_config.csv")
    config = build_holiday_calendar_config(2026, holiday_df, "周末停产", 0.1, 0.1, False, None, None, 0.05, False, None, None, 0.1, [])
    table = normalize_monthly_tou_table(
        pd.DataFrame({"月份": ["2026-01"], "总电量": [1000], "尖电量": [100], "峰电量": [200], "平电量": [300], "谷电量": [400]})
    )
    result, detail = calculate_monthly_tou_consumption(table, 500, 100, holiday_calendar_config=config)
    assert detail.loc[0, "月份"] == 1
    assert detail.loc[0, "节假日修正系数"] < 1
    assert result.post_adjust_self_use_ratio <= result.pre_adjust_self_use_ratio


def test_bill_parser_tou_sum_and_manual_review():
    data = BytesIO()
    pd.DataFrame({"尖电量": [10], "峰电量": [20], "平电量": [30], "谷电量": [40], "金额": [9999]}).to_excel(data, index=False)
    data.seek(0)
    data.name = "bill.xlsx"
    parsed = parse_excel_bill(data)
    assert parsed.annual_consumption_kwh == 100
    assert parsed.extracted_fields["total_kwh_source"] == "tou_sum"

    bad = BytesIO()
    pd.DataFrame({"金额": [9999]}).to_excel(bad, index=False)
    bad.seek(0)
    bad.name = "bad.xlsx"
    parsed_bad = parse_excel_bill(bad)
    assert parsed_bad.parsed_status == "need_manual_review"
    assert parsed_bad.annual_consumption_kwh == 0


def test_revenue_scope_and_tax_note():
    result = calculate_revenue(
        capacity_kwp=100,
        total_investment_yuan=300000,
        equivalent_hours=1000,
        first_year_degradation_ratio=0.02,
        annual_degradation_ratio=0.004,
        operating_years=25,
        self_use_ratio=0.8,
        feed_in_ratio=0.2,
        self_use_tariff=0.8,
        feed_in_tariff=0.4,
        contract_discount_ratio=0.9,
        om_cost_yuan_per_wp_year=0.04,
        insurance_ratio=0.002,
        cleaning_cost_yuan_per_wp_year=0.01,
        roof_rent_yuan_per_year=0,
        other_cost_yuan_per_year=0,
        income_tax_ratio=0.25,
        discount_rate=0.06,
    )
    assert result.finance_model_scope == "simplified_unlevered_preliminary"
    assert "未考虑折旧抵税" in result.tax_simplification_note


def test_export_does_not_mutate_cost_and_writes_detail_sheet():
    cost = {"total": 100, "cost_items_yuan": {"module": 60, "inverter": 40}}
    original = {"total": 100, "cost_items_yuan": {"module": 60, "inverter": 40}}
    sensitivity = pd.DataFrame({"变量": ["投资"], "变化幅度": [0.1], "IRR": [0.1234], "NPV(元)": [1000]})
    data = build_excel_report(
        {"project": "x"},
        {"capacity": 1},
        {"self_use_ratio": 0.8},
        cost,
        {"收益测算口径说明": "简化全投资现金流口径"},
        pd.DataFrame({"year": [1]}),
        sensitivity,
        ["risk"],
        pd.DataFrame({"日期": ["2026-01-01"]}),
        {"module": 0.8, "inverter": 0.1},
    )
    assert cost == original
    workbook = load_workbook(BytesIO(data))
    assert "核心结论摘要" in workbook.sheetnames
    assert "容量测算" in workbook.sheetnames
    assert "消纳测算" in workbook.sheetnames
    assert "造价分项明细" in workbook.sheetnames
    assert "节假日与停产修正" in workbook.sheetnames
    assert "敏感性分析" in workbook.sheetnames
    sensitivity_sheet = workbook["敏感性分析"]
    assert sensitivity_sheet["B2"].number_format == "0.00%"
    assert sensitivity_sheet["C2"].number_format == "0.00%"


if __name__ == "__main__":
    test_capacity_area_basis()
    test_normalize_month_value_and_table()
    test_scale_area_differs_by_pixel_length()
    test_precise_curve_resample_and_tolerance()
    test_monthly_tou_calendar_merge_with_text_month()
    test_bill_parser_tou_sum_and_manual_review()
    test_revenue_scope_and_tax_note()
    test_export_does_not_mutate_cost_and_writes_detail_sheet()
    print("review fixes ok")
